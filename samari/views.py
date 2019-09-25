from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .forms import RSVPForm
from django.core.mail import send_mail
import telepot
from .email_template import rsvp_confirmation, telegram_confirmation, get_message
from datetime import datetime
import os
from .models import RSVP
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import *

def handler404(request, exception, template_name='404.html'):
    response = render(request, template_name)
    response.status_code = 404
    return response

def handler500(request, template_name='500.html'):
    response = render(request, template_name)
    response.status_code = 500
    return response

def index(request):
    if request.method == 'POST':
        request.session['form'] = request.POST
        return redirect('rsvp')
    else:
        form = RSVPForm()

    context = {'form': form, 'scroll': request.method == 'POST'}
    context['dating'] = ['img/gallery/dating/{}.jpg'.format(i + 1) for i in range(12)]
    context['engage'] = ['img/gallery/engage/{}.jpg'.format(i + 1) for i in range(24)]

    return render(request, 'index.html', context)

def rsvp(request):
    if request.method == 'POST' or request.session.get('form'):
        form = RSVPForm(request.POST or request.session.get('form'))
        if request.session.get('form'):
            del request.session['form']
        if form.is_valid():
            data = form.cleaned_data
            html = rsvp_confirmation \
                  .replace('{date}', datetime.now().strftime('%b %d, %Y')) \
                  .replace('{name}', data['name']) \
                  .replace('{message}', get_message(data['attending'] != 'None'))
            send_mail('RSVP Confirmation for Smith-Henry Wedding',
                      'Thank you! We\'ve recorded your response.',
                      'samuel.e.henry97@gmail.com',
                      [data['email']],
                      fail_silently=True,
                      html_message=html)
            sendTelegram(telegram_confirmation.format(data['name'], data['email'], data['number'], attending[data['attending']], data['message']))
            rsvp = form.save()
            messages.success(request, 'Thank you! Your response has been recorded. We\'ll send a confirmation email shortly.')
            return redirect('home')
        else:
            messages.warning(request, 'Please fix the errors below.')
    else:
        form = RSVPForm()

    return render(request, 'rsvp.html', {'form': form})

@login_required
def guests(request):
    fields = ['name', 'email', 'number', 'attending', 'message', 'created']
    widths = [len(field) for field in fields]

    rsvps = RSVP.objects.all().order_by('created')
    total = len(rsvps)

    book = Workbook()
    sheet = book.active
    sheet.title = 'Guests'

    headingFont = Font(bold=True, size=14)
    grayFill = PatternFill(start_color='FFCCCCCC', end_color='FFCCCCCC', fill_type='solid')
    greenFill = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')

    for i, field in enumerate(fields):
        cell = sheet.cell(row=1, column=i+1)
        cell.value = field.capitalize()
        cell.font = headingFont

    for i, rsvp in enumerate(rsvps):
        for j, field in enumerate(fields):
            value = getattr(rsvp, field)
            widths[j] = max(widths[j], len(str(value)))
            sheet.cell(row=i+2, column=j+1).value = value

    for i in range(len(fields)):
        sheet.cell(row=total+2, column=i+1).fill = grayFill

    start = total + 3
    for i, head in enumerate(['Attending', 'All', 'RSVP']):
        cell = sheet.cell(row=start, column=i+1)
        cell.value = head
        cell.font = headingFont

    for i, status in enumerate(['Both', 'Ceremony', 'None']):
        sheet.cell(row=start+i+1, column=1).value = '{}:'.format(status)
        sheet.cell(row=start+i+1, column=2).value = '=SUMIF(D2:D{0}, "{1}", C2:C{0})'.format(total+1, status)
        sheet.cell(row=start+i+1, column=3).value = '=COUNTIF(D2:D{0}, "{1}")'.format(total+1, status)

    values = ['Total:', '=SUM(C2:C{0})'.format(total+1), '=COUNT(C2:C{0})'.format(total+1)]
    for i, value in enumerate(values):
        cell = sheet.cell(row=start+4, column=i+1)
        cell.value = value
        cell.fill = greenFill

    widths[4] = min(widths[4], 50)
    widths[5] = min(widths[5], 18)
    for i, width in enumerate(widths):
        sheet.column_dimensions[get_column_letter(i+1)].width = max(width, 18)

    response = HttpResponse(content=save_virtual_workbook(book), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Guests_{}.xlsx'.format(datetime.now().strftime('%m_%d_%Y'))
    return response

def sendTelegram(message):
    client.sendMessage('131453030', message)
    client.sendMessage('784263152', message)

attending = {'Both': 'Attending Ceremony and Reception', 'Ceremony': 'Attending Ceremony Only', 'None': 'Not Attending'}
client = telepot.Bot(os.environ['WEDDING_BOT'])
